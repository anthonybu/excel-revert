import sys
from os import listdir, path, makedirs
import openpyxl


def revert_excel(source_file, target_path, verbose=False):
    file_name = path.split(source_file)[1]
    print('Processing file: {}'.format(file_name))
    wb_source = openpyxl.load_workbook(source_file)
    ws_source = wb_source.worksheets[0]
    wb_target = openpyxl.Workbook()
    ws_target = wb_target.active
    ws_target.title = 'reverted'

    if verbose:
        print('Total rows in sheet: ' + str(ws_source.max_row))
        print('Total cols in sheet: ' + str(ws_source.max_column))

    for row in range(ws_source.max_row, 0, -1):
        ws_target.append([cell.value for cell in ws_source[row]])

    wb_target.save(path.join(target_path, file_name))


source_path = sys.argv[1:]
source_files = []

if source_path:
    source_path = source_path[0]
    if path.isfile(source_path):
        source_files.append(source_path)
        source_path = path.dirname(source_path)
    else:
        source_files.extend([path.join(source_path, f) for f in listdir(source_path)
                             if path.isfile(path.join(source_path, f))])
else:
    print('Which file(s) to convert?\nYou can drag a file or a folder to the program.')

if source_files:
    print('Found {} file(s) to be processed in: {}\n'.format(len(source_files), source_path))
    choice = input("Would you like to covert these files to reversed order? [Y/N]").lower()
    if choice in ('yes', 'y', ''):
        output_path = source_path + '\\reverted'
        if not path.exists(output_path):
            makedirs(output_path)
        for f in source_files:
            revert_excel(f, output_path)
    else:
        print('Operation aborted by user.')

input('\n\nPress Enter to exit')
